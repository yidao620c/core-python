#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
Topic: 先查询出每天的，然后导出为excel格式
Desc :
"""

import sys
import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.writer.dump_worksheet import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors, borders, fills

import logging
import datetime
import mysql.connector
from mysql.connector import errorcode

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
_log = logging.getLogger('app.' + __name__)

SQL_SELECT_GZ = """
SELECT COUNT(*) FROM t_policy WHERE created_time BETWEEN %s AND %s;
"""

def _connect():
    config = {
        'user': 'root',
        'password': 'mysql',
        'host': '192.168.200.33',
        'database': 'fastloan3',
        'raise_on_warnings': True,
    }
    cnx = None
    try:
        cnx = mysql.connector.connect(**config)
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)
        if cnx:
            cnx.close()
    return cnx


def load_gz_data():
    _log.info('开始从数据库中加载数据')
    format1 = "%Y-%m-%d"
    format2 = "%Y-%m-%d %H:%M:%S"
    start_time = datetime.datetime(2015, 7, 28, 0, 0, 0)
    end_time = datetime.datetime(2015, 7, 28, 23, 59, 59)
    today = datetime.datetime.now()
    detlta_day = datetime.timedelta(days=1)

    conn = _connect()
    cur = conn.cursor()

    result = []
    while start_time < today:
        cur.execute(SQL_SELECT_GZ, (start_time.strftime(format2), end_time.strftime(format2)))
        result.append((start_time.strftime(format1), cur.fetchone()[0]))
        start_time += detlta_day
        end_time += detlta_day
    cur.close()
    conn.close()
    _log.info('数据库中加载数据完毕')
    return result


def export_to_excel(db_data, xlsx_name):
    """导出到excel文件中"""
    _log.info('开始导出到excel文件中')
    border = Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        right=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')
    )
    alignment = Alignment(horizontal='justify',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=True,
                          indent=0)
    fill = PatternFill(fill_type=None, start_color='FFFFFFFF')
    # 基本的样式
    basic_style = Style(font=Font(name='Microsoft YaHei')
                        , border=border, alignment=alignment
                        , fill=fill)
    # header_style = basic_style.copy(
    #     font=Font(name='Microsoft YaHei', b=True, size=15, color='00215757'),
    #     fill=PatternFill(fill_type=fills.FILL_SOLID, start_color='00BAA87F'))
    header_style = basic_style.copy()
    common_style = basic_style.copy()
    wb = Workbook()
    ws = wb.create_sheet(index=0, title='enterprises-{}'.format(len(db_data)))

    ws['A1'] = '日期'
    ws['A1'].style = header_style
    ws['B1'] = '新增数量'
    ws['B1'].style = header_style

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    for i, row in enumerate(db_data):
        ws['A{}'.format(i + 2)] = row[0]
        ws['A{}'.format(i + 2)].style = common_style
        ws['B{}'.format(i + 2)] = row[1]
        ws['B{}'.format(i + 2)].style = common_style
    wb.save(filename=xlsx_name)
    _log.info('导出excel文件完成')


if __name__ == '__main__':
    gz_data = load_gz_data()
    export_to_excel(gz_data, r'D:\download\20160302\新装客户端.xlsx')
    pass
